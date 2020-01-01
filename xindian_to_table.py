# coding: utf-8
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.styles import Font
import operator
import os
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.cell import WriteOnlyCell

G_C_M = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']  # 列标题名字M性别
G_C_F = ['N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y']  # 列标题名字F性别

def id_to_no(id, animal_noid):  # 根据ID返回NO
    for i in animal_noid:
        if id.strip()[1] != 'F' and id.strip()[1] != 'M':
            if id.strip() == str(i[1]):
                return str(i[0])
        if id.strip()[1] == 'F' or id.strip()[1] == 'M':
            return id.strip()
    return '异常ID-{}'.format(id.strip())
    


def no_to_id(id, animal_noid):  # 根据NO返回ID
    for i in animal_noid:
        if id.strip()[1] == 'F' or id.strip()[1] == 'M':
            if id.strip() == str(i[0]):
                return str(i[1])
        if id.strip()[1] != 'F' and id.strip()[1] != 'M':
            return id.strip()
    return '异常ID-{}'.format(id.strip())


def read_xindian_xlsx(path_xlsx_list):

    fenzu_fname = path_xlsx_list + '\\' + '分组表.xls'
    df_study = pd.read_csv(open(fenzu_fname), nrows=1)
    # study_name = str(df_study.columns[0]).split('for ')[1]
    df_fenzu = pd.read_csv(open(fenzu_fname), skiprows=2)
    df_fenzu.reset_index()
    animal_no = df_fenzu['Study animal number'].tolist()
    animal_id = df_fenzu['Pretest number'].tolist()
    animal_noid = list(zip(animal_no, animal_id))
    xlsx_list = [fn for fn in os.listdir(path_xlsx_list) if fn.endswith('.xlsx')]
    xlsx_list.sort()
    xlsx_L_name = list(map(lambda y: path_xlsx_list + '\\' + y, xlsx_list))
    wb = load_workbook(xlsx_L_name[0])
    wb.guess_types = False
    ws = wb.active
    ws_data_M = []
    ws_data_F = []
    '''
    columns_list = ['动物ID', '动物编号', '检测时间', '心率(次/分)',
                    'P-R间期(s)', 'QRS间期(s)', 'Q-T间期(s)', 'P波(mV)',
                    'QRS波(mV)', 'T波(mV)', 'ST段(mV)']
    '''
    '''求excel 有数据内容最大长度以E列最后一个由数据的值为准'''
    sheet_lenM = ws.max_row
    try:
        sheet_lenM = [i for i in range(3, len(list(ws.iter_rows(max_col=1))) + 3) if not ws['E{}'.format(i)].value][0]
    except:
        pass
    sheet_lenF = ws.max_row
    try:
        sheet_lenF = [i for i in range(3, len(list(ws.iter_rows(max_col=1))) + 3) if not ws['R{}'.format(i)].value][0]
    except:
        pass

    '''选取数据组成list为DataFarme准备.分性别MF'''
    for i in range(3, sheet_lenM, 3):  # 3-最大内容长度,步进3
        for x in range(3):  # 输入三次数据
            ws_data_M.append([ws.cell(row=i, column=1).value,
                              id_to_no(str(ws.cell(row=i, column=1).value), animal_noid),
                              ws.cell(row=i, column=3).value,
                              ws.cell(row=i, column=4).value])
    for i in range(3, sheet_lenF, 3):  # 3-最大内容长度,步进3
        for x in range(3):  # 输入三次数据
            ws_data_F.append([ws.cell(row=i, column=14).value,
                              id_to_no(str(ws.cell(row=i, column=14).value), animal_noid),
                              ws.cell(row=i, column=16).value,
                              ws.cell(row=i, column=17).value])
    for x, i in enumerate(range(3, sheet_lenM)):  # 步进1获取数据组成列表M
        for m in range(5, 12):
            ws_data_M[x].append(ws.cell(row=i, column=m).value)
    for x, i in enumerate(range(3, sheet_lenF)):  # 步进1获取数据组成列表F
        for f in range(18, 25):
            ws_data_F[x].append(ws.cell(row=i, column=f).value)
    M_title = ws.cell(row=1, column=1).value
    F_title = ws.cell(row=1, column=14).value

    '''试验阶段排序'''
    '''性别M'''
    study_PDR_M = [i[2] for i in ws_data_M]
    PDR_set = list(set(study_PDR_M))
    PDR_set.sort(key=study_PDR_M.index)  # 保留顺序转换列表->去重复集合->列表原始顺序排序
    Data_sort_M = []
    for x in PDR_set:
        x_list = []
        for i in ws_data_M:
            if x == i[2]:
                x_list.append(i)
        x_list.sort(key=operator.itemgetter(1))
        Data_sort_M = Data_sort_M + x_list

    '''性别F'''
    study_PDR_F = [i[2] for i in ws_data_F]
    PDR_set = list(set(study_PDR_F))
    PDR_set.sort(key=study_PDR_F.index)  # 保留顺序转换列表->去重复集合->列表原始顺序排序
    Data_sort_F = []
    for x in PDR_set:
        x_list = []
        for i in ws_data_F:
            if x == i[2]:
                x_list.append(i)
        x_list.sort(key=operator.itemgetter(1))
        Data_sort_F = Data_sort_F + x_list

    return Data_sort_M, Data_sort_F, M_title, F_title


def to_xls_xindian(to_xls_name, ws_data_M, ws_data_F, M_title, F_title):
    wb = Workbook()  # 实例化
    ws = wb.active  # 激活一个sheet 默认0 sheet
    Table_Title_M = M_title
    Table_Title_F = F_title
    Title_2 = ['动物ID', '动物编号', '检测时间', '心率(次/分)',
               'P-R间期(s)', 'QRS间期(s)', 'Q-T间期(s)', 'P波(mV)',
               'QRS波(mV)', 'T波(mV)', 'ST段(mV)', 'ST段(mV)绝对值']
    Title_2_width = 10.50
    Title_2_height = 30.00
    Alig_center = Alignment(horizontal='center', vertical='center',
                            wrapText=True)  # 双剧中
    '''单元格线框颜色设定'''
    border_All = Border(left=Side(border_style='thin', color='FF000000'),
                        right=Side(border_style='thin', color='FF000000'),
                        top=Side(border_style='thin', color='FF000000'),
                        bottom=Side(border_style='thin', color='FF000000'),
                        )
    font_A = Font(name='Times New Roman', size=11)
    font_B = Font(name='Times New Roman', size=11, bold=True)

    '''合并单元格,并填充数据Title'''
    ws['A1'].font = Font(name='Times New Roman', size=22, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    ws.cell(row=1, column=1, value=Table_Title_M)

    ws['N1'].font = Font(name='Times New Roman', size=22, bold=True)
    ws['N1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=1, start_column=14, end_row=1, end_column=25)
    ws.cell(row=1, column=14, value=Table_Title_F)

    '''设定行高列宽'''
    for i in G_C_M + G_C_F:
        ws.column_dimensions[i].width = Title_2_width
        ws2_s = ws['{}2'.format(i)]
        ws2_s.font = Font(name='Times New Roman', size=11, bold=True)
        # ws2_s.font = Font(name=u'宋体')
        ws2_s.alignment = Alignment(horizontal='center', vertical='center',
                                    wrapText=True)  # 设定自动换行

    for i in range(1, 3):  # 第1,2行高
        ws.row_dimensions[i].height = Title_2_height

    '''填充第2行数据'''
    for i in range(1, 13):
        ws.cell(row=2, column=i, value=Title_2[i-1])
        ws.cell(row=2, column=i,).border = border_All
        ws.cell(row=2, column=i + 13, value=Title_2[i-1])
        ws.cell(row=2, column=i + 13).border = border_All

    '''数据操作填充操作'''
    # start_num = 3  # 开始行编号
    to_len_xlsxM = int(len(ws_data_M) / 3 * 4) + 3  # 应该写入到地excel长度M
    to_len_xlsxF = int(len(ws_data_F) / 3 * 4) + 3  # 应该写入到地excel长度F
    slM = sorted([i for i in range(3, to_len_xlsxM, 4)] +
                 [i for i in range(4, to_len_xlsxM, 4)] +
                 [i for i in range(5, to_len_xlsxM, 4)])  # 生成一个数据列表
    slF = sorted([i for i in range(3, to_len_xlsxF, 4)] +
                 [i for i in range(4, to_len_xlsxF, 4)] +
                 [i for i in range(5, to_len_xlsxF, 4)])  # 生成一个数据列表
    tlM = [i for i in range(6, to_len_xlsxM, 4)]  # 生成用于计算的列表列
    tlF = [i for i in range(6, to_len_xlsxF, 4)]  # 生成用于计算的列表列
    for n, i in enumerate(slM):
        for l, m in enumerate(range(1, 12)):  # 代表1-12
            ws.cell(row=i, column=m, value=ws_data_M[n][l])
            ws.cell(row=i, column=m).alignment = Alig_center
            ws.cell(row=i, column=m).number_format = '0.00'
        ws.cell(row=i, column=1).number_format = '0'
        ws['L{}'.format(i)] = "=ABS(K{})".format(i)
        ws.cell(row=i, column=12).alignment = Alig_center
        ws.cell(row=i, column=12).number_format = '0.00'
    for n, i in enumerate(slF):
        for l, f in enumerate(range(14, 25)):  # 代表14-25
            ws.cell(row=i, column=f, value=ws_data_F[n][l])
            ws.cell(row=i, column=f).alignment = Alig_center
            ws.cell(row=i, column=f).number_format = '0.00'
        ws.cell(row=i, column=14).number_format = '0'
        ws['Y{}'.format(i)] = "=ABS(X{})".format(i)
        ws.cell(row=i, column=25).alignment = Alig_center
        ws.cell(row=i, column=25).number_format = '0.00'
    for i in slM:
        for x in range(1, 26):
            ws.cell(row=i, column=x).font = font_A
    for i in slF:
        for x in range(1, 26):
            ws.cell(row=i, column=x).font = font_A

    '''公式数据列'''
    GC_MFM = ['A', 'B', 'D']  # 等于公式列
    GC_MFF = ['N', 'O', 'Q']
    mean_lM = ['C']  # mean列
    mean_lF = ['P']
    GC_M = ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']  # 均值公式列表
    GC_F = ['R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y']  # 均值公式列表
    for i in tlM:  # 均值公式行
        for mean in mean_lM:  # mean列
            ws['{}{}'.format(mean, i)] = "mean"
        for z in GC_MFM:  # 等值公式列
            ws['{}{}'.format(z, i)] = "={}{}".format(z, i - 1)
        for m in GC_M:  # 均值公式列M
            ws['{}{}'.format(m, i)] = "=AVERAGE({}{}:{}{})".format(m, i - 3, m, i - 1)

    for i in tlF:  # 均值公式行
        for mean in mean_lF:  # mean列
            ws['{}{}'.format(mean, i)] = "mean"
        for z in GC_MFF:  # 等值公式列
            ws['{}{}'.format(z, i)] = "={}{}".format(z, i - 1)
        for f in GC_F:  # 均值公式列F
            ws['{}{}'.format(f, i)] = "=AVERAGE({}{}:{}{})".format(f, i - 3, f, i - 1)

    '''加线框居中设定公式行加粗'''
    for i in range(3, to_len_xlsxM):
        for x in range(1, 26):  # 所有行列
            ws.cell(row=i, column=x).alignment = Alig_center
            ws.cell(row=i, column=x).border = border_All
        ws.cell(row=i, column=4).number_format = '0'
        ws.cell(row=i, column=17).number_format = '0'
    for i in tlM:
        for x in range(1, 5):
            ws.cell(row=i, column=x).font = font_B          
        for x in range(14, 18):
            ws.cell(row=i, column=x).font = font_B
        for x in range(5, 13):  # M计算数值列
            ws.cell(row=i, column=x).number_format = '0.00'
            ws.cell(row=i, column=x).font = font_B
        for x in range(18, 26):  # F计算数值列
            ws.cell(row=i, column=x).number_format = '0.00'
            ws.cell(row=i, column=x).font = font_B
    
    '''统计分析导出表格部分'''
    '''性别M转换到一个文件'''
    s_spss_M = []
    s_spss_F = []
    for x, row in enumerate(range(6, to_len_xlsxM, 4)):  # 从第6行开始步进4获取数据形成列表
        '''性别M获取数据'''
        A_No_v1 = str(ws.cell(row=row - 3, column=2).value)
        Col_3_V = str(ws.cell(row=row - 3, column=3).value)
        Col_4_V = str(ws.cell(row=row - 3, column=4).value)
        s_spss_M.append([A_No_v1, Col_3_V, Col_4_V])
        for y in [5, 6, 7, 8, 9, 10, 11]:
            '''E-K列均值  5-11'''
            mean_l = np.mean([ws.cell(row=row - 3, column=y).value,
                              ws.cell(row=row - 2, column=y).value,
                              ws.cell(row=row - 1, column=y).value])
            s_spss_M[x].append('{:.2f}'.format(mean_l))
        '''l列绝对值  12 '''
        abs_l = np.mean([abs(ws.cell(row=row - 3, column=11).value),
                         abs(ws.cell(row=row - 2, column=11).value),
                         abs(ws.cell(row=row - 1, column=11).value)])
        s_spss_M[x].append('{:.2f}'.format(abs_l))
    for x, row in enumerate(range(6, to_len_xlsxF, 4)):  # 从第6行开始步进4获取数据形成列表
        '''性别F获取数据'''
        A_No_v2 = str(ws.cell(row=row - 3, column=15).value)
        Col_16_V = str(ws.cell(row=row - 3, column=16).value)
        Col_17_V = str(ws.cell(row=row - 3, column=17).value)
        s_spss_F.append([A_No_v2, Col_16_V, Col_17_V])
        for y in [18, 19, 20, 21, 22, 23, 24]:
            '''R-X列均值  18-24'''
            mean_l = np.mean([ws.cell(row=row - 3, column=y).value,
                              ws.cell(row=row - 2, column=y).value,
                              ws.cell(row=row - 1, column=y).value])
            s_spss_F[x].append('{:.2f}'.format(mean_l))
        '''Y列绝对值  25 '''
        abs_l = np.mean([abs(ws.cell(row=row - 3, column=24).value),
                         abs(ws.cell(row=row - 2, column=24).value),
                         abs(ws.cell(row=row - 1, column=24).value)])
        s_spss_F[x].append('{:.2f}'.format(abs_l))

    df2 = spss_to_df(s_spss_F)
    df1 = spss_to_df(s_spss_M)

    ws1 = wb.create_sheet(title='心电spssF', index=3)
    for r in dataframe_to_rows(df1, index=False, header=True):
        ws1.append(r)

    ws2 = wb.create_sheet(title='心电spssF', index=3)
    for r in dataframe_to_rows(df2, index=False, header=True):
        ws2.append(r)
    wb.save(to_xls_name)


def spss_to_df(s_spss):  # 输入list 转换为pd.DataFrame
    df_spss = pd.DataFrame(s_spss, columns=['动物编号',
                                            '试验阶段',
                                            '01_心率(次/分)',
                                            '02_P-R间期(s)',
                                            '03_QRS间期(s)',
                                            '04_Q-T间期(s)',
                                            '05_P波(mV)',
                                            '06_QRS波(mV)',
                                            '07_T波(mV)',
                                            '08_ST段(mV)',
                                            '09_ST段(mV)绝对值'])

    '''转换前试验阶段排序相关的准备工作'''
    df_spss = df_spss[~df_spss['动物编号'].isin(['-'])]  # 反选删除没用的动物编号-
    df_spss_2list = df_spss.iloc[:, 1].tolist()
    PDR_set = list(set(df_spss_2list))
    PDR_set.sort(key=df_spss_2list.index)  # 保留顺序转换列表->去重复集合->列表原始顺序排序
    PDR_set_l = []
    for x, i in enumerate(PDR_set, 1):
        PDR_set_l.append((i, '{:02d}_{}'.format(x, i)))

    def id_to_idn(id, id_list):  # 根据及生成的列表加01- 用于排序
        for id_old in id_list:
            if id == id_old[0]:
                return id_old[1]

    df_spss['试验阶段'] = df_spss[:]['试验阶段'].map(lambda x: id_to_idn(str(x), PDR_set_l))
    # df_spss['Group'] = df_spss[:]['动物编号'].map(lambda x: str(x).strip()[:1])
    # df_spss = df_spss.set_index('Group', '试验阶段', '动物编号',)
    df_spss.index = [df_spss['动物编号'].tolist(), df_spss['试验阶段'].tolist()]
    df_spss.drop(['动物编号'], axis=1, inplace=True)
    df_spss.drop(['试验阶段'], axis=1, inplace=True)
    df_spss = df_spss.unstack()  # 长列转宽列列名
    CO1_name = list(df_spss.columns.levels[0])
    CO2_name = list(df_spss.columns.levels[1])
    RO2_name = list(df_spss.index)

    '''组合成一个列名字'''
    title_list = list(
        map(lambda x: '{}'.format(str(x), str(x)), [str(i) + '_' + str(j)[3:] for i in CO1_name for j in CO2_name]))
    df_spss_ok = pd.DataFrame(df_spss.values, columns=title_list, index=RO2_name)
    df_spss_ok = df_spss_ok.reset_index()
    df_spss_ok.rename(columns={'index': 'ANIMAL_NO'}, inplace=True)  # 重名命列名字
    df_spss_ok.insert(0, 'Group', df_spss_ok[:]['ANIMAL_NO'].map(lambda x: str(x).strip()[:1]))  # 插入一列group
    # print(df_spss_ok)
    return df_spss_ok


def xin_dian(xlsx_path, to_xls_name):
    M_l, F_l, M_title, F_title = read_xindian_xlsx(xlsx_path)
    to_xls_xindian(to_xls_name + '\\' + '心电.xlsx', M_l, F_l, M_title, F_title)


if __name__ == "__main__":
    path_csv_list = r'C:\Users\admin\Desktop\22'
    xin_dian(path_csv_list, path_csv_list)
    print('ok')