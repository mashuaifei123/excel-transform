# coding: utf-8
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.styles import Font
import os
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.cell import WriteOnlyCell

G_C_M = ['A', 'B', 'C', 'D', 'E', 'F']  # 列标题名字M性别
G_C_F = ['H', 'I', 'J', 'K', 'L', 'M']  # 列标题名字F性别


def read_csv(path_csv_list):
    '''从一个文件夹路径获取csv文件.拼接返回一个DF, 实验名字, 对照表'''
    csv_list = [fn for fn in os.listdir(path_csv_list) if fn.lower().endswith('.csv')]
    print(csv_list)
    csv_list.sort()
    csv_L_name = list(map(lambda y: path_csv_list + '\\' + y, csv_list))
    fenzu_fname = path_csv_list + '\\' + '分组表.xls'
    df_study = pd.read_csv(open(fenzu_fname), nrows=1)
    study_name = str(df_study.columns[0]).split('for ')[1]
    df_fenzu = pd.read_csv(open(fenzu_fname), skiprows=2)
    df_fenzu.reset_index()
    animal_no = df_fenzu['Study animal number'].tolist()
    animal_id = df_fenzu['Pretest number'].tolist()
    animal_noid = list(zip(animal_no, animal_id))
    df_all = pd.read_csv(open(csv_L_name[0]))
    if len(csv_L_name) > 1:
        for i in range(1, len(csv_L_name)):
            df = pd.read_csv(open(csv_L_name[i]))
            df_all = pd.concat([df_all, df], axis=0)  # 拼接
    df_all = df_all.reset_index()
    # df_all.to_csv('sss.csv', encoding='GBK')
    return df_all, animal_noid, study_name


def retuen_jjz(b):
    '''
    返回最接近值所在单元格
    in:[(<Cell 'Sheet'.F7>, 117), (<Cell 'Sheet'.F6>, 119), (<Cell 'Sheet'.F5>,
         122)]
    out:(<Cell 'Sheet'.F7>, <Cell 'Sheet'.F6>)
    '''
    b = [i[0] for i in b]
    a = sorted(b, key=lambda x: x[1])
    if abs(a[0][1]-a[1][1]) < abs(a[1][1]-a[2][1]):
        return a[0][0], a[1][0]
    else:
        return a[1][0], a[2][0]


def id_to_no(id, animal_noid):  # 根据ID返回NO
    for i in animal_noid:
        if id.strip()[1] != 'F' and id.strip()[1] != 'M':
            if id.strip() == str(i[1]):
                return str(i[0])
        if id.strip()[1] == 'F' or id.strip()[1] == 'M':
            return id.strip()
    return '异常ID-{}'.format(id.strip())
    


def no_to_id(id, animal_noid):  # 根据NO返回ID
    for i in animal_noid:
        if id.strip()[1] == 'F' or id.strip()[1] == 'M':
            if id.strip() == str(i[0]):
                return str(i[1])
        if id.strip()[1] != 'F' and id.strip()[1] != 'M':
            return id.strip()
    return '异常ID-{}'.format(id.strip())
    


def df_to_xls_xueya(to_xls_name, Data_Frame, animal_noid, study_name='study_name'):

    error_list = []
    df = Data_Frame
    wb = Workbook()  # 实例化
    ws = wb.active  # 激活一个sheet 默认0 sheet
    Table_Title_M = '{} 雄性动物血压'.format(study_name)
    Table_Title_F = '{} 雌性动物血压'.format(study_name)
    Title_2 = ['动物ID', '动物编号', '试验阶段', 'SBP（mmHg）',
               'DBP（mmHg）', 'MBP（mmHg）']
    Title_2_width = 10.50
    Title_2_height = 30.00
    Rows_height = 15.00
    Alig_center = Alignment(horizontal='center', vertical='center',
                            wrapText=True)  # 双剧中
    '''单元格线框颜色设定'''
    border_All = Border(left=Side(border_style='thin', color='FF000000'),
                        right=Side(border_style='thin', color='FF000000'),
                        top=Side(border_style='thin', color='FF000000'),
                        bottom=Side(border_style='thin', color='FF000000'),
                        )
    font_B = Font(name='Times New Roman', size=11, bold=True)

    '''合并单元格,并填充数据Title'''
    ws['A1'].font = Font(name='Times New Roman', size=22, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    ws.cell(row=1, column=1, value=Table_Title_M)

    ws['H1'].font = Font(name='Times New Roman', size=22, bold=True)
    ws['H1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=1, start_column=8, end_row=1, end_column=13)
    ws.cell(row=1, column=8, value=Table_Title_F)

    '''设定行高列宽'''
    list_Title_2 = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L',
                    'M']
    for i in list_Title_2:
        ws.column_dimensions[i].width = Title_2_width
        ws2_s = ws['{}2'.format(i)]
        ws2_s.font = Font(name='Times New Roman', size=11, bold=True)
        # ws2_s.font = Font(name=u'宋体')
        ws2_s.alignment = Alignment(horizontal='center', vertical='center',
                                    wrapText=True)  # 设定自动换行

    for i in range(1, 3):  # 第1,2行高
        ws.row_dimensions[i].height = Title_2_height

    '''填充第2行数据'''
    for i in range(1, 7):
        ws.cell(row=2, column=i, value=Title_2[i-1])
        ws.cell(row=2, column=i,).border = border_All
        ws.cell(row=2, column=i + 7, value=Title_2[i-1])
        ws.cell(row=2, column=i + 7).border = border_All

    '''数据操作填充操作'''
    ''' # 可以使用的第一种情况
    df['试验阶段'] = df[:]['试验阶段'].map(lambda x: str(x).split('-')[1])
    '''
    '''
    df['试验阶段'] = df[:]['试验阶段'].map(lambda x: '{}-{}'.format(str(x).split('-')[1],
                                        str(x).split('-')[-1]))
    '''
    df['ANIMAL_NO'] = df[:]['ANIMAL NO'].map(lambda x: id_to_no(str(x), animal_noid))  # 修改动物编号为 no
    df['Animal_ID'] = df[:]['ANIMAL NO'].map(lambda x: no_to_id(str(x), animal_noid))  # 增加一列动物ID
    df.insert(2, 'M_F', df[:]['ANIMAL_NO'].map(lambda x: str(x).strip()[1:2]))  # 插入新的一列性别
    # df.to_csv('df.csv', encoding='GBK')
    df_M = df[df[u'M_F'].isin(['M'])]  # 筛选性别M
    df_M = df_M.sort_values(by=['DATE', 'ANIMAL_NO'])
    # df_M.to_csv('M.csv', encoding='GBK')
    df_F = df[df[u'M_F'].isin(['F'])]  # 筛选性别F
    df_F = df_F.sort_values(by=['DATE', 'ANIMAL_NO'])
    # df_F.to_csv('F.csv', encoding='GBK')
    df_gro_M = df_M.groupby(['DATE', 'ANIMAL_NO'])  # 分组
    df_gro_F = df_F.groupby(['DATE', 'ANIMAL_NO'])  # 分组
    start_num_M = 3  # 开始行编号M
    start_num_F = 3  # 开始行编号F
    com_text = '注：mean为标*两次数据的平均值。'

    for k, v in df_gro_M.groups:  # 性别M分组后的组名字
        '''性别M导入excel'''
        cv_l = [(1, 15), (2, 14), (3, 8), (4, 11), (5, 12), (6, 13)]
        df_groups_i = df_gro_M.get_group((k, v))  # 根据组名返回的DataFrame
        df_groups_i_size = df_groups_i.index.size  # DataFrame 行长度
        if df_groups_i_size == 1:  # 只有一条数据情况
            for cvl in cv_l:
                ws.cell(row=start_num_M, column=cvl[0], value=df_groups_i.iloc[0][cvl[1]])
                ws.cell(row=start_num_M, column=cvl[0],).border = border_All
            start_num_M += 1

        elif df_groups_i_size == 3:  # 只有三条数据情况
            for i in range(0, df_groups_i_size):
                for cvl in cv_l:
                    ws.cell(row=start_num_M, column=cvl[0], value=df_groups_i.iloc[i][cvl[1]])
                    ws.cell(row=start_num_M, column=cvl[0],).border = border_All
                start_num_M += 1

            '''增加一行为计算平均值切加粗行'''
            ws.cell(row=start_num_M, column=1, value=df_groups_i.iloc[0][15])
            ws.cell(row=start_num_M, column=2, value=df_groups_i.iloc[0][14])
            ws.cell(row=start_num_M, column=3, value='mean')

            '''根据MBP值近似值返回cell位置数设定平均值公式'''
            dict_Index_col_value = []
            for row in ws.iter_rows(min_row=start_num_M - 3, min_col=6, max_col=6, max_row=start_num_M - 1):  # 也可以使用openpyxl.worksheet.Worksheet.iter_rows()这个方法
                dict_Index_col_value.append([(cell, cell.value) for cell in row])
            cell_num = retuen_jjz(dict_Index_col_value)
            cn_1 = int(str(cell_num[0]).split('F')[1][:-1])
            cn_2 = int(str(cell_num[1]).split('F')[1][:-1])
            ws.cell(row=cn_1, column=3, value='{}*'.format(ws['C{}'.format(cn_1)].value))
            ws.cell(row=cn_2, column=3, value='{}*'.format(ws['C{}'.format(cn_2)].value))
            ws['D{}'.format(start_num_M)] = "=AVERAGE(D{},D{})".format(cn_1, cn_2)
            ws.cell(row=start_num_M, column=4).number_format = '0'
            ws['E{}'.format(start_num_M)] = "=AVERAGE(E{},E{})".format(cn_1, cn_2)
            ws.cell(row=start_num_M, column=5).number_format = '0'
            ws['F{}'.format(start_num_M)] = "=AVERAGE(F{},F{})".format(cn_1, cn_2)
            ws.cell(row=start_num_M, column=6).number_format = '0'

            for cvl in cv_l:
                ws.cell(row=start_num_M, column=cvl[0],).border = border_All
                ws.cell(row=start_num_M, column=cvl[0],).font = font_B
            start_num_M += 1

        else:
            print('发现异常数据:', k, v)
            error_list.append((k, v))
            pass

    for k, v in df_gro_F.groups:  # 性别F分组后的组名字
        '''性别F导入excel'''
        cv_l = [(8, 15), (9, 14), (10, 8), (11, 11), (12, 12), (13, 13)]
        df_groups_i = df_gro_F.get_group((k, v))  # 根据组名返回的DataFrame
        df_groups_i_size = df_groups_i.index.size  # DataFrame 行长度
        if df_groups_i_size == 1:
            for cvl in cv_l:
                ws.cell(row=start_num_F, column=cvl[0], value=df_groups_i.iloc[0][cvl[1]])
                ws.cell(row=start_num_F, column=cvl[0],).border = border_All
            start_num_F += 1

        elif df_groups_i_size == 3:
            for i in range(0, df_groups_i_size):
                for cvl in cv_l:
                    ws.cell(row=start_num_F, column=cvl[0], value=df_groups_i.iloc[i][cvl[1]])
                    ws.cell(row=start_num_F, column=cvl[0],).border = border_All
                start_num_F += 1

            '''增加一行为计算平均值切加粗行'''
            ws.cell(row=start_num_F, column=8, value=df_groups_i.iloc[0][15])
            ws.cell(row=start_num_F, column=9, value=df_groups_i.iloc[0][14])
            ws.cell(row=start_num_F, column=10, value='mean')

            '''根据MBP值近似值返回cell位置数设定平均值公式'''
            dict_Index_col_value = []
            for row in ws.iter_rows(min_row=start_num_F - 3, min_col=13, max_col=13, max_row=start_num_F - 1):  # 也可以使用openpyxl.worksheet.Worksheet.iter_rows()这个方法
                dict_Index_col_value.append([(cell, cell.value) for cell in row])
            cell_num = retuen_jjz(dict_Index_col_value)
            cn_1 = int(str(cell_num[0]).split('M')[1][:-1])
            cn_2 = int(str(cell_num[1]).split('M')[1][:-1])
            ws.cell(row=cn_1, column=10, value='{}*'.format(ws['J{}'.format(cn_1)].value))
            ws.cell(row=cn_2, column=10, value='{}*'.format(ws['J{}'.format(cn_2)].value))
            ws['K{}'.format(start_num_F)] = "=AVERAGE(K{},K{})".format(cn_1, cn_2)
            ws.cell(row=start_num_F, column=11).number_format = '0'
            ws['L{}'.format(start_num_F)] = "=AVERAGE(L{},L{})".format(cn_1, cn_2)
            ws.cell(row=start_num_F, column=12).number_format = '0'
            ws['M{}'.format(start_num_F)] = "=AVERAGE(M{},M{})".format(cn_1, cn_2)
            ws.cell(row=start_num_F, column=13).number_format = '0'
            for cvl in cv_l:
                ws.cell(row=start_num_F, column=cvl[0],).font = font_B
                ws.cell(row=start_num_F, column=cvl[0],).border = border_All
            start_num_F += 1

        else:
            print('发现异常数据:', k, v)
            error_list.append((k, v))
            pass

    '''最后添加合并行'''
    ws['A{}'.format(start_num_M)].font = Font(name='Times New Roman', size=11)
    ws.cell(row=start_num_M, column=1, value=com_text)
    ws['H{}'.format(start_num_F)].font = Font(name='Times New Roman', size=11)
    ws.cell(row=start_num_F, column=8, value=com_text)
    ws.merge_cells(start_row=start_num_M, start_column=1, end_row=start_num_M, end_column=6)
    ws.merge_cells(start_row=start_num_F, start_column=8, end_row=start_num_F, end_column=13)

    for i in range(1, 7):
        ws.cell(row=start_num_M, column=i,).border = border_All
    for i in range(8, 14):
        ws.cell(row=start_num_F, column=i,).border = border_All

    for i in ws.rows:  # 设定所有居中显示
        for k in i:
            k.alignment = Alig_center

    if error_list:  # 判断增加异常数据显示
        ws.cell(row=start_num_M + 3, column=1, value='异常数据:{}'.format(str(error_list)))


    '''统计分析导出表格部分'''
    '''性别M转换到一个文件'''
    s_spss_M = []
    for row in range(3, start_num_M):
        A_No_v1 = ws.cell(row=row, column=2).value
        A_No_v2 = ws.cell(row=row + 1, column=2).value
        Col_3_V = ws.cell(row=row, column=3).value
        Col_3_V2 = str(ws.cell(row=row - 1, column=3).value).replace('*', '')
        Col_4_V = ws.cell(row=row, column=4).value
        Col_5_V = ws.cell(row=row, column=5).value
        Col_6_V = ws.cell(row=row, column=6).value
        if A_No_v1 != A_No_v2:  # 动物ID不等于下一列的
            if Col_3_V == 'mean':  # 均值情况需要计算
                '''F列均值'''
                cellF0 = str(ws['F{}'.format(row)].value).split('(')[1].split(',')
                cellFA = cellF0[0]
                cellFB = cellF0[1][:-1]
                cell_mean_F = (int(ws[cellFA].value) + int(ws[cellFB].value)) / 2
                '''E列均值'''
                cellE0 = str(ws['E{}'.format(row)].value).split('(')[1].split(',')
                cellEA = cellE0[0]
                cellEB = cellE0[1][:-1]
                cell_mean_E = (int(ws[cellEA].value) + int(ws[cellEB].value)) / 2
                '''D列均值'''
                cellD0 = str(ws['D{}'.format(row)].value).split('(')[1].split(',')
                cellDA = cellD0[0]
                cellDB = cellD0[1][:-1]
                cell_mean_D = (int(ws[cellDA].value) + int(ws[cellDB].value)) / 2
                s_spss_M.append([A_No_v1[0], A_No_v1, Col_3_V2, cell_mean_D, cell_mean_E,
                              cell_mean_F])
            else:  # 正常数据对的情况
                s_spss_M.append([A_No_v1[0], A_No_v1, Col_3_V, Col_4_V, Col_5_V,
                              Col_6_V])
    df1 = spss_to_df(s_spss_M)
    ws1 = wb.create_sheet(title='血压spssM', index=2)
    for r in dataframe_to_rows(df1,index=False, header=True):
        ws1.append(r)


    '''性别F转换到一个文件'''
    s_spss_F = []
    for row in range(3, start_num_F):
        A_No_v1 = ws.cell(row=row, column=9).value
        A_No_v2 = ws.cell(row=row + 1, column=9).value
        Col_3_V = ws.cell(row=row, column=10).value
        Col_3_V2 = str(ws.cell(row=row - 1, column=10).value).replace('*', '')
        Col_4_V = ws.cell(row=row, column=11).value
        Col_5_V = ws.cell(row=row, column=12).value
        Col_6_V = ws.cell(row=row, column=13).value
        if A_No_v1 != A_No_v2:  # 动物ID不等于下一列的
            if Col_3_V == 'mean':  # 均值情况需要计算
                '''M列均值'''
                cellM0 = str(ws['M{}'.format(row)].value).split('(')[1].split(',')
                cellMA = cellM0[0]
                cellMB = cellM0[1][:-1]
                cell_mean_M = (int(ws[cellMA].value) + int(ws[cellMB].value)) / 2
                '''L列均值'''
                cellL0 = str(ws['L{}'.format(row)].value).split('(')[1].split(',')
                cellLA = cellL0[0]
                cellLB = cellL0[1][:-1]
                cell_mean_L = (int(ws[cellLA].value) + int(ws[cellLB].value)) / 2
                '''K列均值'''
                cellK0 = str(ws['K{}'.format(row)].value).split('(')[1].split(',')
                cellKA = cellK0[0]
                cellKB = cellK0[1][:-1]
                cell_mean_K = (int(ws[cellKA].value) + int(ws[cellKB].value)) / 2
                s_spss_F.append([A_No_v1[0], A_No_v1, Col_3_V2, cell_mean_K, cell_mean_L,
                                cell_mean_M])
            else:  # 正常数据对的情况
                s_spss_F.append([A_No_v1[0], A_No_v1, Col_3_V, Col_4_V, Col_5_V,
                                Col_6_V])

    df2 = spss_to_df(s_spss_F)
    ws2 = wb.create_sheet(title='血压spssF', index=3)
    for r in dataframe_to_rows(df2, index=False, header=True):
        ws2.append(r)
    wb.save(to_xls_name)

def spss_to_df(s_spss):  # 输入list 转换为pd.DataFrame
    df_spss = pd.DataFrame(s_spss, columns=['Group', '动物编号', '试验阶段',
                                            'SBP（mmHg）', 'DBP（mmHg）', 'MBP（mmHg）'])
    Title1_spss = ['Group', '动物编号', '试验阶段', 'SBP（mmHg）', 'DBP（mmHg）',
                    'MBP（mmHg）']
    '''转换前试验阶段排序相关的准备工作'''
    df_spss_2list = df_spss.iloc[:, 2].tolist()
    PDR_set = list(set(df_spss_2list))
    PDR_set.sort(key=df_spss_2list.index)  # 保留顺序转换列表->去重复集合->列表原始顺序排序
    PDR_set_l = []
    for x, i in enumerate(PDR_set, 1):
        PDR_set_l.append((i, '{:02d}-{}'.format(x, i)))

    def id_to_idn(id, id_list):  # 根据及生成的列表加01- 用于排序
        for id_old in id_list:
            if id == id_old[0]:
                return id_old[1]
    df_spss['试验阶段'] = df_spss[:]['试验阶段'].map(lambda x: id_to_idn(str(x), PDR_set_l))
    df_spss = df_spss.pivot_table(index=['Group', '动物编号'], columns='试验阶段')  # 数据透视
    # df_spss = df_spss.reset_index()  # 索引重排例
    CO1_name = list(df_spss.columns.levels[0])
    CO2_name = list(df_spss.columns.levels[1])
    RO2_name = list(df_spss.index.levels[1])
    '''组合成一个列名字'''
    title_list = list(map(lambda x: '{}'.format(str(x), str(x)), [str(i) + '_' + str(j)[3:] for i in CO1_name for j in CO2_name]))
    df_spss_ok = pd.DataFrame(df_spss.values, columns=title_list, index=RO2_name)
    df_spss_ok = df_spss_ok.reset_index()
    df_spss_ok.rename(columns={'index': 'ANIMAL_NO'}, inplace=True)  # 重名命列名字
    df_spss_ok.insert(0, 'Group', df_spss_ok[:]['ANIMAL_NO'].map(lambda x: str(x).strip()[:1]))  # 插入一列group
    return df_spss_ok


def xue_ya(csv_path, to_xls_name):
    df_all, animal_noid, study_name = read_csv(csv_path)
    df_to_xls_xueya(to_xls_name + '\\' + '血压.xlsx', df_all, animal_noid, study_name)


if __name__ == "__main__":
    path_csv_list = r'C:\Users\admin\Desktop\22'
    xue_ya(path_csv_list, path_csv_list)
    