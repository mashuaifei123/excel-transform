# import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.cell import WriteOnlyCell

def read_ex(path_xls):
    df = pd.read_excel(path_xls)
    return df


def df_to_xls(Data_Frame_list, to_xls_name, Table_Title_A='A2018025-T014-01'):
    df = Data_Frame_list[0]
    wb = Workbook()  # 实例化
    ws = wb.active  # 激活一个sheet 默认0 sheet
    Table_Title_M = '{} 雄性动物呼吸'.format(Table_Title_A)
    Table_Title_F = '{} 雌性动物呼吸'.format(Table_Title_A)
    Title_2 = ['动物ID', '动物编号', '检测时间', '呼吸频率\n（次/分）',
               '最大峰值（g）', '最小谷值（g）', '呼吸幅度（g）']
    Title_2_width = 10.50
    Title_2_height = 30.00
    Rows_height = 15.00
    Alig_center = Alignment(horizontal='center', vertical='center',
                            wrapText=True)  # 双剧中
    '''单元格线框颜色设定'''
    border_All = Border(left=Side(border_style='thin', color='FF000000'),
                        right=Side(border_style='thin', color='FF000000'),
                        top=Side(border_style='thin', color='FF000000'),
                        bottom=Side(border_style='thin', color='FF000000'),
                        )

    '''合并单元格,并填充数据Title'''
    ws['A1'].font = Font(name='Times New Roman', size=22, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    ws.cell(row=1, column=1, value=Table_Title_M)

    ws['I1'].font = Font(name='Times New Roman', size=22, bold=True)
    ws['I1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=1, start_column=9, end_row=1, end_column=15)
    ws.cell(row=1, column=9, value=Table_Title_F)

    '''设定行高列宽'''
    list_Title_2 = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L',
                    'M', 'N', 'O']
    for i in list_Title_2:
        ws.column_dimensions[i].width = Title_2_width
        ws2_s = ws['{}2'.format(i)]
        ws2_s.font = Font(name='Times New Roman', size=11, bold=True)
        # ws2_s.font = Font(name=u'宋体')
        ws2_s.alignment = Alignment(horizontal='center', vertical='center',
                                    wrapText=True)  # 设定自动换行

    for i in range(1, 3):  # 第1,2行高
        ws.row_dimensions[i].height = Title_2_height

    '''填充第2行数据'''
    for i in range(1, 8):
        ws.cell(row=2, column=i, value=Title_2[i-1])
        ws.cell(row=2, column=i + 8, value=Title_2[i-1])

    '''数据操作填充操作'''
    # 可以使用的第一种情况
    df['试验阶段'] = df[:]['试验阶段'].map(lambda x: str(x).split('-')[-1])
    # df['试验阶段'] = df[:]['试验阶段'].map(lambda x: '{}-{}'.format(str(x).split('-')[1],str(x).split('-')[-1]))
    df.insert(2, 'M_F', df[:]['动物编号'].map(lambda x: str(x)[1:2]))  # 插入新的一列性别
    df_M = df[df[u'M_F'].isin(['M'])]  # 筛选性别M
    df_M.reset_index()
    df_F = df[df[u'M_F'].isin(['F'])]  # 筛选性别F
    df_F.reset_index()

    for n, i in enumerate(range(3, df_M.index.size * 4 + 3, 4)):
        '''性别M'''
        ws.cell(row=i, column=1, value=df_M.iloc[n][-1])
        ws.cell(row=i, column=2, value=df_M.iloc[n][0])
        ws.cell(row=i, column=3, value=df_M.iloc[n][1])
        ws.cell(row=i, column=4, value=df_M.iloc[n][6])
        ws.cell(row=i, column=5, value=df_M.iloc[n][3])
        ws.cell(row=i, column=6, value=df_M.iloc[n][4])
        ws['G{}'.format(i)] = "=E{}-F{}".format(i, i)

        ws.cell(row=i+1, column=1, value=df_M.iloc[n][-1])
        ws.cell(row=i+1, column=2, value=df_M.iloc[n][0])
        ws.cell(row=i+1, column=3, value=df_M.iloc[n][1])
        ws.cell(row=i+1, column=4, value=df_M.iloc[n][10])
        ws.cell(row=i+1, column=5, value=df_M.iloc[n][7])
        ws.cell(row=i+1, column=6, value=df_M.iloc[n][8])
        ws['G{}'.format(i+1)] = "=E{}-F{}".format(i+1, i+1)

        ws.cell(row=i+2, column=1, value=df_M.iloc[n][-1])
        ws.cell(row=i+2, column=2, value=df_M.iloc[n][0])
        ws.cell(row=i+2, column=3, value=df_M.iloc[n][1])
        ws.cell(row=i+2, column=4, value=df_M.iloc[n][14])
        ws.cell(row=i+2, column=5, value=df_M.iloc[n][11])
        ws.cell(row=i+2, column=6, value=df_M.iloc[n][12])
        ws['G{}'.format(i+2)] = "=E{}-F{}".format(i+2, i+2)

        ws.cell(row=i+3, column=1, value=df_M.iloc[n][-1])
        ws.cell(row=i+3, column=2, value=df_M.iloc[n][0])
        ws.cell(row=i+3, column=3, value='mean')
        ws['D{}'.format(i+3)] = "=AVERAGE(D{}:D{})".format(i, i+2)
        ws['E{}'.format(i+3)] = "=AVERAGE(E{}:E{})".format(i, i+2)
        ws['F{}'.format(i+3)] = "=AVERAGE(F{}:F{})".format(i, i+2)
        ws['G{}'.format(i+3)] = "=AVERAGE(G{}:G{})".format(i, i+2)

        font_B = Font(name='Times New Roman', size=11, bold=True)
        ws['A{}'.format(i+3)].font = font_B
        ws['B{}'.format(i+3)].font = font_B
        ws['C{}'.format(i+3)].font = font_B
        ws['D{}'.format(i+3)].font = font_B
        ws['E{}'.format(i+3)].font = font_B
        ws['F{}'.format(i+3)].font = font_B
        ws['G{}'.format(i+3)].font = font_B

    for n, i in enumerate(range(3, df_F.index.size * 4 + 3, 4)):
        '''性别F'''
        ws.cell(row=i, column=9, value=df_F.iloc[n][-1])
        ws.cell(row=i, column=10, value=df_F.iloc[n][0])
        ws.cell(row=i, column=11, value=df_F.iloc[n][1])
        ws.cell(row=i, column=12, value=df_F.iloc[n][6])
        ws.cell(row=i, column=13, value=df_F.iloc[n][3])
        ws.cell(row=i, column=14, value=df_F.iloc[n][4])
        ws['O{}'.format(i)] = "=M{}-N{}".format(i, i)

        ws.cell(row=i+1, column=9, value=df_F.iloc[n][-1])
        ws.cell(row=i+1, column=10, value=df_F.iloc[n][0])
        ws.cell(row=i+1, column=11, value=df_F.iloc[n][1])
        ws.cell(row=i+1, column=12, value=df_F.iloc[n][10])
        ws.cell(row=i+1, column=13, value=df_F.iloc[n][7])
        ws.cell(row=i+1, column=14, value=df_F.iloc[n][8])
        ws['O{}'.format(i+1)] = "=M{}-N{}".format(i+1, i+1)

        ws.cell(row=i+2, column=9, value=df_F.iloc[n][-1])
        ws.cell(row=i+2, column=10, value=df_F.iloc[n][0])
        ws.cell(row=i+2, column=11, value=df_F.iloc[n][1])
        ws.cell(row=i+2, column=12, value=df_F.iloc[n][14])
        ws.cell(row=i+2, column=13, value=df_F.iloc[n][11])
        ws.cell(row=i+2, column=14, value=df_F.iloc[n][12])
        ws['O{}'.format(i+2)] = "=M{}-N{}".format(i+2, i+2)

        ws.cell(row=i+3, column=9, value=df_F.iloc[n][-1])
        ws.cell(row=i+3, column=10, value=df_F.iloc[n][0])
        ws.cell(row=i+3, column=11, value='mean')
        ws['L{}'.format(i+3)] = "=AVERAGE(L{}:L{})".format(i, i+2)
        ws['M{}'.format(i+3)] = "=AVERAGE(M{}:M{})".format(i, i+2)
        ws['N{}'.format(i+3)] = "=AVERAGE(N{}:N{})".format(i, i+2)
        ws['O{}'.format(i+3)] = "=AVERAGE(O{}:O{})".format(i, i+2)
        font_B = Font(name='Times New Roman', size=11, bold=True)
        ws['I{}'.format(i+3)].font = font_B
        ws['J{}'.format(i+3)].font = font_B
        ws['K{}'.format(i+3)].font = font_B
        ws['L{}'.format(i+3)].font = font_B
        ws['M{}'.format(i+3)].font = font_B
        ws['N{}'.format(i+3)].font = font_B
        ws['O{}'.format(i+3)].font = font_B

    '''设定剧中对其'''
    for i in list_Title_2:
        for o in range(3, df_M.index.size * 4 + 3):
            ws['{}{}'.format(i, o)].alignment = Alig_center

    for i in list_Title_2:
        for o in range(3, df_F.index.size * 4 + 3):
            ws['{}{}'.format(i, o)].alignment = Alig_center

    '''设定单元格数值显示格式'''
    for i in range(3, df_M.index.size * 4 + 3):
        '''性别M'''
        ws.cell(row=i, column=4).number_format = '0'
        ws.cell(row=i, column=5).number_format = '0.0'
        ws.cell(row=i, column=6).number_format = '0.0'
        ws.cell(row=i, column=7).number_format = '0.0'
    for i in range(3, df_F.index.size * 4 + 3):
        '''性别F'''
        ws.cell(row=i, column=12).number_format = '0'
        ws.cell(row=i, column=13).number_format = '0.0'
        ws.cell(row=i, column=14).number_format = '0.0'
        ws.cell(row=i, column=15).number_format = '0.0'

    '''设定边框线条'''
    for i in range(2, df_M.index.size * 4 + 3):  # M性别
        for m in range(1, 8):
            ws.cell(row=i, column=m).border = border_All

    for i in range(2, df_F.index.size * 4 + 3):  # F性别
        for f in range(9, 16):
                ws.cell(row=i, column=f).border = border_All

    df1 = Data_Frame_list[1]
    ws1 = wb.create_sheet(title='呼吸幅度M', index=2)
    for r in dataframe_to_rows(df1, index=True, header=True):
        ws1.append(r)
    df2 = Data_Frame_list[2]
    ws1 = wb.create_sheet(title='呼吸幅度F', index=3)
    for r in dataframe_to_rows(df2, index=True, header=True):
        ws1.append(r)
    df3 = Data_Frame_list[3]
    ws1 = wb.create_sheet(title='呼吸频率M', index=4)
    for r in dataframe_to_rows(df3, index=True, header=True):
        ws1.append(r)
    df4 = Data_Frame_list[4]
    ws1 = wb.create_sheet(title='呼吸频率F', index=5)
    for r in dataframe_to_rows(df4, index=True, header=True):
        ws1.append(r)

    wb.save(to_xls_name)


if __name__ == "__main__":
    df_to_xls(read_ex('12332.xlsx'), 'new2.xlsx')
